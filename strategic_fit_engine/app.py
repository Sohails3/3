"""
Strategic Fit Engine — Web Interface

Serves a form where the user inputs buyer, sector, and geography.
Runs the 4-step pipeline in a background thread and streams
real-time progress to the browser via Server-Sent Events (SSE).
When complete, the report opens automatically.

Run:
    cd "/Users/sohail/Documents/Claude VS"
    ANTHROPIC_API_KEY=sk-ant-... .venv/bin/python -m strategic_fit_engine.app

Open: http://localhost:5000
"""
from __future__ import annotations

import json
import os
import secrets
import sys
import threading
import time
import traceback
import uuid
from pathlib import Path

import sys
# Ensure the workspace root is on the path so `from strategic_fit_engine import ...` works
_ROOT = str(Path(__file__).parent.parent)
if _ROOT not in sys.path:
    sys.path.insert(0, _ROOT)

from dotenv import load_dotenv
from flask import Flask, Response, request, send_file, session, redirect, url_for

load_dotenv(Path(__file__).parent.parent / ".env")

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", secrets.token_hex(32))


def _h(v: str) -> str:
    """HTML-escape a value."""
    return (str(v)
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', "&quot;"))
BASE = Path(__file__).parent

# GP Bullhound bull logo (inline SVG, white version for dark backgrounds)
# ---------------------------------------------------------------------------
# Multi-user job state — keyed by browser session ID
# ---------------------------------------------------------------------------

_jobs: dict = {}          # session_id → job dict
_jobs_lock = threading.Lock()

# Password gate — set ACCESS_PASSWORD env var to require a password
ACCESS_PASSWORD: str = os.environ.get("ACCESS_PASSWORD", "")


class _Capture:
    """Redirects print() calls from step modules into the job message list."""
    def __init__(self, session_id: str) -> None:
        self._sid = session_id
    def write(self, text: str) -> None:
        if text.strip():
            _add(self._sid, "log", text.strip())
    def flush(self) -> None:
        pass


def _add(session_id: str, msg_type: str, text: str, step: int = None) -> None:
    with _jobs_lock:
        job = _jobs.get(session_id)
        if job is None:
            return
        if step is not None:
            job["current_step"] = step
        job["messages"].append({
            "type": msg_type,
            "text": text,
            "step": job["current_step"],
        })


def _run_pipeline(company: str, sector: str, geography: str,
                  mode: str, session_id: str) -> None:
    """Runs all 4 steps sequentially in a background thread."""
    _env_path = Path(__file__).parent.parent / ".env"
    if _env_path.exists():
        for line in _env_path.read_text().splitlines():
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, _, v = line.partition("=")
                if k.strip() not in os.environ:
                    os.environ[k.strip()] = v.strip()

    is_sell = (mode == "sell")
    output_dir = BASE / "output" / session_id
    scored_path = BASE / "data" / f"{session_id}_scored.json"

    old_stdout = sys.stdout
    sys.stdout = _Capture(session_id)
    try:
        _add(session_id, "info", f"Pipeline starting ({mode}-side): {company} · {sector} · {geography}")

        (BASE / "data").mkdir(exist_ok=True)
        output_dir.mkdir(parents=True, exist_ok=True)

        # ── Step 1 ──────────────────────────────────────────────────────
        if is_sell:
            _add(session_id, "step", "Step 1 / 4 — Seller Profile & Acquirer Criteria", step=1)
            from strategic_fit_engine import step1_seller_dna
            bp = step1_seller_dna.run(company, sector)
            n_crit = len(bp.get("scoring_criteria", []))
            _add(session_id, "done", f"✓ Step 1 complete — {company} seller profile built, {n_crit} acquirer criteria derived")
        else:
            _add(session_id, "step", "Step 1 / 4 — Buyer Strategy, Dry Powder & Market Intelligence", step=1)
            from strategic_fit_engine import step1_buyer_dna
            bp = step1_buyer_dna.run(company, sector)
            n_bacq = len(bp.get("buyer_acquisitions", []))
            n_crit = len(bp.get("scoring_criteria", []))
            _add(session_id, "done", f"✓ Step 1 complete — {company} M&A thesis built, {n_bacq} prior acquisitions mapped, {n_crit} scoring criteria derived")
        if bp.get("target_brief"):
            _add(session_id, "log", f"   Brief: {bp['target_brief'][:120]}…")

        # ── Step 2 ──────────────────────────────────────────────────────
        if is_sell:
            _add(session_id, "step", "Step 2 / 4 — Potential Acquirer Longlist", step=2)
            from strategic_fit_engine import step2_acquirer_discovery
            tr = step2_acquirer_discovery.run(sector, geography, buyer_profile=bp)
            n_tgt = len(tr.get("targets", []))
            _add(session_id, "done", f"✓ Step 2 complete — {n_tgt} potential acquirers identified")
        else:
            _add(session_id, "step", "Step 2 / 4 — Buyer-Led Target Longlist", step=2)
            from strategic_fit_engine import step2_discovery
            tr = step2_discovery.run(sector, geography, buyer_profile=bp)
            n_tgt = len(tr.get("targets", []))
            _add(session_id, "done", f"✓ Step 2 complete — {n_tgt} companies on longlist")
        for t in tr.get("targets", [])[:5]:
            _add(session_id, "log", f"   · {t['name']} ({t.get('country','')}) — {t.get('funding_stage','')}")
        if n_tgt > 5:
            _add(session_id, "log", f"   · ... and {n_tgt - 5} more")

        # ── Step 3 ──────────────────────────────────────────────────────
        step3_label = "Acquirer Fit Scoring (8 criteria)" if is_sell else "Strategic Fit Scoring (8 criteria, 2 batches)"
        _add(session_id, "step", f"Step 3 / 4 — {step3_label}", step=3)
        from strategic_fit_engine import step3_scoring
        scored = step3_scoring.run(bp, tr)
        scored["sector"] = sector
        scored["geography"] = geography
        scored["mode"] = mode
        with open(scored_path, "w", encoding="utf-8") as f:
            json.dump(scored, f, indent=2, ensure_ascii=False)
        top = scored["targets"][0] if scored.get("targets") else {}
        top_label = "Top acquirer" if is_sell else "Top target"
        _add(session_id, "done", f"✓ Step 3 complete — {top_label}: {top.get('name','?')} ({top.get('total_score','?')}/{top.get('max_score','?')})")
        for t in scored.get("targets", [])[:3]:
            _add(session_id, "log", f"   #{t['rank']}: {t['name']} — {t['total_score']}/{t['max_score']}")

        # ── Step 4 ──────────────────────────────────────────────────────
        _add(session_id, "step", "Step 4 / 4 — Generating Interactive Dashboard", step=4)
        from strategic_fit_engine import step4_output
        step4_output.run(bp, scored, output_dir=output_dir)
        _add(session_id, "done", "✓ Step 4 complete — Dashboard saved")

        with _jobs_lock:
            if session_id in _jobs:
                _jobs[session_id]["done"] = True
        _add(session_id, "complete", "Analysis complete! Your report is ready.")

    except Exception as e:
        tb = traceback.format_exc()
        _add(session_id, "error", f"Error: {e}")
        _add(session_id, "error", tb)
        with _jobs_lock:
            if session_id in _jobs:
                _jobs[session_id]["error"] = str(e)
    finally:
        sys.stdout = old_stdout
        with _jobs_lock:
            if session_id in _jobs:
                _jobs[session_id]["running"] = False


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    return LANDING_HTML


@app.route("/run", methods=["POST"])
def run_analysis():
    # Assign a persistent session ID for this browser
    sid = session.get("sid")
    if not sid:
        sid = str(uuid.uuid4())
        session["sid"] = sid

    company   = request.form.get("company", "").strip() or request.form.get("buyer", "").strip()
    sector    = request.form.get("sector", "").strip()
    geography = request.form.get("geography", "").strip()
    mode      = request.form.get("mode", "buy").strip()
    if mode not in ("buy", "sell"):
        mode = "buy"
    if not company or not sector or not geography:
        return "Missing required fields: company, sector, geography", 400

    with _jobs_lock:
        job = _jobs.get(sid)
        if job and job["running"]:
            return (
                "<html><body style='font-family:sans-serif;padding:40px'>"
                "<h2>Analysis already running</h2>"
                "<p>An analysis is currently in progress. Wait a few minutes and "
                "<a href='/'>go back</a> to check, or force-reset below.</p>"
                "<form method='post' action='/reset'>"
                "<button type='submit' style='background:#CC0605;color:#fff;border:none;"
                "padding:10px 20px;font-size:14px;cursor:pointer'>Force Reset</button>"
                "</form></body></html>"
            ), 409
        _jobs[sid] = {"running": True, "messages": [], "current_step": 0, "done": False, "error": None}

    t = threading.Thread(target=_run_pipeline,
                         args=(company, sector, geography, mode, sid), daemon=True)
    t.start()
    return _progress_html(company, sector, geography, mode)


@app.route("/stream")
def stream():
    sid = session.get("sid")

    def generate():
        if not sid or sid not in _jobs:
            yield f"data: {json.dumps({'type': 'end', 'error': None})}\n\n"
            return
        last = 0
        while True:
            with _jobs_lock:
                job   = _jobs.get(sid, {})
                msgs  = list(job.get("messages", []))
                done  = job.get("done", False)
                error = job.get("error")

            for msg in msgs[last:]:
                yield f"data: {json.dumps(msg)}\n\n"
            last = len(msgs)

            if done or error:
                yield f"data: {json.dumps({'type': 'end', 'error': error})}\n\n"
                break
            time.sleep(0.25)

    resp = Response(generate(), mimetype="text/event-stream")
    resp.headers["Cache-Control"]      = "no-cache"
    resp.headers["X-Accel-Buffering"]  = "no"
    resp.headers["Connection"]         = "keep-alive"
    return resp


@app.route("/report")
def view_report():
    sid = session.get("sid")
    if sid:
        path = BASE / "output" / sid / "report.html"
        if path.exists():
            return send_file(str(path.resolve()))
    # Legacy fallback (single-user path)
    path = BASE / "output" / "report.html"
    if path.exists():
        return send_file(str(path.resolve()))
    return "Report not yet generated. Run an analysis first.", 404


@app.route("/reset", methods=["POST"])
def reset_job():
    sid = session.get("sid")
    if sid:
        with _jobs_lock:
            _jobs[sid] = {"running": False, "messages": [], "current_step": 0, "done": False, "error": None}
    return "", 204


@app.route("/download/pptx")
def download_pptx():
    sid = session.get("sid")
    path = (BASE / "output" / sid / "report.pptx") if sid else None
    if path is None or not path.exists():
        path = BASE / "output" / "report.pptx"
    if path.exists():
        return send_file(str(path.resolve()),
                         as_attachment=True,
                         download_name="MA_Target_Screening.pptx",
                         mimetype="application/vnd.openxmlformats-officedocument.presentationml.presentation")
    return "PowerPoint not yet generated. Run an analysis first.", 404


@app.route("/download/excel")
def download_excel():
    import json, io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    sid = session.get("sid")
    scored_path = (BASE / "data" / f"{sid}_scored.json") if sid else None
    if scored_path is None or not scored_path.exists():
        scored_path = BASE / "data" / "targets_scored.json"
    if not scored_path.exists():
        return "No screening data found. Run an analysis first.", 404

    with open(scored_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    targets  = data.get("targets", [])
    buyer    = data.get("buyer", "")
    sector   = data.get("sector", "")
    geo      = data.get("geography", "")
    is_sell  = (data.get("mode") == "sell")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Acquirer Screening" if is_sell else "Target Screening"

    # Styles
    navy_fill   = PatternFill("solid", fgColor="0B1C3D")
    gold_fill   = PatternFill("solid", fgColor="C9A84C")
    green_fill  = PatternFill("solid", fgColor="E8F5E9")
    amber_fill  = PatternFill("solid", fgColor="FFF8E1")
    grey_fill   = PatternFill("solid", fgColor="F5F5F5")
    white_font  = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
    bold_font   = Font(bold=True, name="Calibri", size=10)
    normal_font = Font(name="Calibri", size=10)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left        = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin        = Side(style="thin", color="CCCCCC")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title row
    ws.merge_cells("A1:N1")
    title_cell = ws["A1"]
    sheet_title = "Exit Strategy — Acquirer Screening" if is_sell else "M&A Target Screening"
    title_cell.value = f"{sheet_title} — {buyer} | {sector} | {geo}"
    title_cell.font  = white_font
    title_cell.fill  = navy_fill
    title_cell.alignment = center
    ws.row_dimensions[1].height = 28

    # Header row
    headers = [
        "Rank", "Company", "Country", "Founded", "Acquirer Type" if is_sell else "Stage",
        "Revenue (€M)" if is_sell else "ARR (€M)",
        "Fund Size / Mkt Cap (€M)" if is_sell else "Raised (€M)",
        "Headcount" if is_sell else "Employees",
        "Total Score", "Strategic Fit Summary",
        "Key Customers / Portfolio", "Key Investors / LPs",
        "Appetite Score" if is_sell else "Readiness Score",
        "Appetite Summary" if is_sell else "Readiness Summary",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font      = white_font
        cell.fill      = gold_fill
        cell.alignment = center
        cell.border    = border
    ws.row_dimensions[2].height = 22

    # Data rows
    for row_idx, t in enumerate(targets, start=3):
        rank  = t.get("rank", row_idx - 2)
        score = t.get("total_score", 0)
        maxs  = t.get("max_score", 40)

        if rank <= 3:
            fill = green_fill
        elif rank <= 6:
            fill = amber_fill
        else:
            fill = grey_fill

        arr_val    = t.get("arr_usd_m", "N/A")
        raised_val = t.get("total_raised_usd_m", "N/A")
        try:    arr_val    = round(float(arr_val), 1)
        except: arr_val    = "N/A"
        try:    raised_val = round(float(raised_val), 1)
        except: raised_val = "N/A"

        def _ls(v):
            if isinstance(v, list):
                parts = [x for x in v if x and x != "Not publicly available"]
                return ", ".join(parts) if parts else "N/A"
            return str(v) if v else "N/A"

        values = [
            rank,
            t.get("name", ""),
            t.get("country", ""),
            t.get("founded", ""),
            t.get("funding_stage", ""),
            arr_val,
            raised_val,
            t.get("employees", "N/A"),
            f"{score}/{maxs}",
            t.get("strategic_fit_summary", t.get("product_description", "")),
            _ls(t.get("key_customers", [])),
            _ls(t.get("key_investors", [])),
            t.get("readiness_score", ""),
            t.get("readiness_summary", ""),
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill      = fill
            cell.font      = bold_font if col in (1, 2, 9) else normal_font
            cell.alignment = center if col in (1, 3, 4, 6, 7, 8, 9, 13) else left
            cell.border    = border
        ws.row_dimensions[row_idx].height = 40

    # Column widths
    col_widths = [6, 22, 14, 10, 16, 10, 12, 12, 12, 50, 30, 30, 14, 40]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="Exit_Acquirer_Screening.xlsx" if is_sell else "MA_Target_Screening.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/download/workflow-pptx")
def download_workflow_pptx():
    from strategic_fit_engine import workflow_pptx
    path = workflow_pptx.run()
    return send_file(str(path.resolve()),
                     as_attachment=True,
                     download_name="GP_Bullhound_Workflow_Diagrams.pptx",
                     mimetype="application/vnd.openxmlformats-officedocument.presentationml.presentation")


@app.route("/workflow")
def view_workflow():
    path = BASE / "output" / "workflow_diagrams.html"
    if path.exists():
        return send_file(str(path.resolve()))
    return "Workflow diagrams not found.", 404


# ---------------------------------------------------------------------------
# Password gate
# ---------------------------------------------------------------------------

@app.before_request
def require_auth():
    """Redirect to login if ACCESS_PASSWORD is set and user is not authenticated."""
    if not ACCESS_PASSWORD:
        return  # No password configured — open access
    if request.endpoint in ("login", "auth"):
        return  # These routes handle auth themselves
    if not session.get("auth"):
        return redirect(url_for("login"))


@app.route("/login", methods=["GET"])
def login():
    return LOGIN_HTML


@app.route("/auth", methods=["POST"])
def auth():
    password = request.form.get("password", "")
    if password == ACCESS_PASSWORD:
        session["auth"] = True
        return redirect(url_for("index"))
    return LOGIN_HTML.replace(
        'id="pw-error" style="display:none"',
        'id="pw-error"'
    ), 401


LOGIN_HTML = """<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>Strategic Fit Engine — Access</title>
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link href="https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@300;400;600;700&display=swap" rel="stylesheet">
  <style>
    *{box-sizing:border-box;margin:0;padding:0}
    body{font-family:'IBM Plex Sans','Helvetica Neue',Arial,sans-serif;
      background:#252850;min-height:100vh;display:flex;flex-direction:column;
      align-items:center;justify-content:center;color:#fff}
    .card{width:100%;max-width:420px;background:#fff;padding:48px 48px 40px;color:#252850}
    .logo{display:flex;align-items:center;gap:12px;margin-bottom:32px}
    .logo-text{font-size:13px;font-weight:700;letter-spacing:.04em;color:#888;
      text-transform:uppercase}
    .logo-text span{color:#CC0605}
    .logo-badge{background:#252850;padding:6px 12px;display:flex;align-items:center;gap:7px}
    .logo-badge-label{font-size:11px;font-weight:700;color:#fff;letter-spacing:.02em;
      font-family:'IBM Plex Sans',sans-serif}
    h1{font-size:26px;font-weight:700;color:#252850;margin-bottom:8px;line-height:1.2}
    .sub{font-size:13px;color:#888;margin-bottom:32px;line-height:1.6}
    label{font-size:10px;font-weight:700;letter-spacing:.1em;text-transform:uppercase;
      color:#252850;display:block;margin-bottom:8px}
    input[type=password]{width:100%;border:1px solid #d0d5e8;padding:13px 14px;
      font-size:14px;color:#252850;font-family:inherit;outline:none;
      transition:border .15s;background:#fff;margin-bottom:20px}
    input[type=password]:focus{border-color:#252850}
    button{width:100%;background:#CC0605;color:#fff;border:none;padding:14px;
      font-size:13px;font-weight:700;font-family:inherit;cursor:pointer;
      letter-spacing:.04em;text-transform:uppercase;transition:background .15s}
    button:hover{background:#a80504}
    #pw-error{color:#CC0605;font-size:12px;margin-top:14px;
      font-weight:600;display:none}
    .footer{margin-top:32px;font-size:11px;color:rgba(255,255,255,.3);text-align:center}
  </style>
</head>
<body>
  <div class="card">
    <div class="logo">
      <div class="logo-badge">
        <svg xmlns="http://www.w3.org/2000/svg" width="83.729" height="65.861" viewBox="0 0 83.729 65.861" style="height:20px;width:auto"><path d="M145.038,197c.1.1.2.207.306.309a.355.355,0,0,0,.5.094c.078-.034.158-.064.231-.094a.7.7,0,0,1,.3.449,1.8,1.8,0,0,0,.137.358c.113.232.245.455.35.689a.8.8,0,0,0,.523.459c.426.134.848.283,1.273.422a4.67,4.67,0,0,1,1.585.844c.574.481,1.142.968,1.713,1.452.3.251.593.5.886.752a1.421,1.421,0,0,1,.54.855c.028.186.012.378.022.566.013.274.016.55.051.822a.971.971,0,0,0,.575.775,2.5,2.5,0,0,0,.28.132,17.342,17.342,0,0,0,2.68.835,3.631,3.631,0,0,1,.695.192,1.027,1.027,0,0,1,.67,1.3c-.038.15-.094.295-.139.443a3.085,3.085,0,0,0-.09,1.253c.026.222.067.442.108.662a1.654,1.654,0,0,1-.025.716,5.265,5.265,0,0,1-.812,1.91,3.657,3.657,0,0,1-1.262,1.123,13.617,13.617,0,0,1-1.348.581,3.1,3.1,0,0,1-.658.123.8.8,0,0,0-.548.306,1.792,1.792,0,0,1-1.5.677,4.88,4.88,0,0,1-2.954-.92,1.292,1.292,0,0,0-.953-.235,3.979,3.979,0,0,0-.851.2,7.019,7.019,0,0,0-3.112,2.349,3.159,3.159,0,0,0-.673,1.944c-.006.774.044,1.549.075,2.322a13.282,13.282,0,0,1-.078,2.5,5.247,5.247,0,0,1-.188.8,19.087,19.087,0,0,1-4.387,7.4c-.554.589-1.135,1.154-1.7,1.729-.164.165-.33.328-.494.492a6.688,6.688,0,0,0-1.014,1.3c-1.725,2.869-3.468,5.728-5.2,8.6a42.319,42.319,0,0,0-2.144,4.148c-.435.944-.844,1.9-1.21,2.876a7.841,7.841,0,0,0-.512,3.01c.017.568-.012,1.137.01,1.7a7.676,7.676,0,0,0,.132,1.308,2.628,2.628,0,0,0,1.277,1.71c.278.171.567.327.837.511a1.388,1.388,0,0,1,.62,1.028,1.674,1.674,0,0,1-1.025,1.71,2.842,2.842,0,0,1-.879.18c-.947.018-1.894.005-2.841-.009a1.229,1.229,0,0,1-.823-.386,2.3,2.3,0,0,1-.556-.89,6.318,6.318,0,0,1-.35-2.1c-.011-.525.022-1.051-.005-1.575-.069-1.307-.157-2.612-.245-3.918-.047-.7-.106-1.39-.171-2.084a2.1,2.1,0,0,0-.126-.6,1.781,1.781,0,0,1,.241-1.78,9.679,9.679,0,0,0,1.594-3.688,8.327,8.327,0,0,0,.025-2.644c-.095-.673-.206-1.345-.263-2.021a10.35,10.35,0,0,1,0-1.471c.04-.678.117-1.355.176-2.032.007-.076,0-.152,0-.247-.162-.019-.322-.043-.482-.056a28.429,28.429,0,0,1-4.219-.712,57.123,57.123,0,0,1-5.665-1.75,21.981,21.981,0,0,0-2.692-.7,25.049,25.049,0,0,0-3.4-.47c-.9-.066-1.8-.133-2.706-.156a15.573,15.573,0,0,0-3.086.216,3.48,3.48,0,0,0-1.049.346,1.673,1.673,0,0,0-.831,1c-.315.964-.613,1.937-.987,2.878a13.815,13.815,0,0,1-3.263,4.908,31.925,31.925,0,0,1-6.736,5.008,10.825,10.825,0,0,1-1.026.484q-1.3.547-2.6,1.072a2.921,2.921,0,0,0-1.085.726,9.845,9.845,0,0,0-.826.991c-.5.718-.994,1.447-1.463,2.189-.45.713-.889,1.435-1.291,2.176-.422.778-.8,1.582-1.188,2.376a2.879,2.879,0,0,0-.149.412,1.709,1.709,0,0,0,.283,1.611c.206.276.435.534.643.809a1.5,1.5,0,0,1,.151,1.518.6.6,0,0,1-.374.372,3.43,3.43,0,0,1-.769.2c-1.081.081-2.165.138-3.248.2a2.546,2.546,0,0,1-.54-.038,1.046,1.046,0,0,1-.827-.628,1.965,1.965,0,0,1-.114-1.553q.158-.423.342-.835c.458-1.029.934-2.051,1.38-3.086q.878-2.041,1.918-4a13.764,13.764,0,0,0,1.27-3.451c.106-.469.265-.925.395-1.389a1.863,1.863,0,0,1,.8-1.041,8.944,8.944,0,0,0,.935-.728,4.048,4.048,0,0,0,1.194-1.895c.8-2.74,1.75-5.434,2.665-8.139.4-1.188.718-2.406,1.072-3.61a.235.235,0,0,0,0-.073c-.26.176-.508.362-.772.521a9.348,9.348,0,0,1-2.234.926,13.335,13.335,0,0,1-4.288.511,8.826,8.826,0,0,1-3.243-.777,3.714,3.714,0,0,1-.976-.623,2.49,2.49,0,0,1-.562-.708,1.759,1.759,0,0,1,.2-.054,7.485,7.485,0,0,1,1.8.029,15.821,15.821,0,0,0,2.239.134,8.979,8.979,0,0,0,7.616-4.409c.495-.8.938-1.625,1.39-2.446a15.758,15.758,0,0,1,3.708-4.5,5.175,5.175,0,0,1,.729-.531,19.466,19.466,0,0,1,5.258-2.153,17.969,17.969,0,0,1,2.6-.394c1.151-.1,2.3-.09,3.458-.086.843,0,1.686-.047,2.529-.094.592-.033,1.184-.09,1.774-.151a17.729,17.729,0,0,0,2.725-.566c1.6-.426,3.2-.875,4.81-1.282,1.475-.373,2.962-.7,4.442-1.054a43.05,43.05,0,0,0,7.265-2.4c.747-.326,1.527-.578,2.293-.863l.629-.231a6.817,6.817,0,0,0,3.239-2.423c.845-1.155,1.756-2.255,2.705-3.324a24.031,24.031,0,0,1,3.589-3.41l.416-.307a1.655,1.655,0,0,0,.6-.822,5.144,5.144,0,0,1,.555-1.19,4.973,4.973,0,0,1,1.624-1.525q.9-.549,1.809-1.1a.813.813,0,0,0,.11-.1Z" transform="translate(-74.671 -197)" fill="#fff"/></svg>
        <span class="logo-badge-label">GP Bullhound</span>
      </div>
      <div class="logo-text">Strategic Fit Engine <span>·</span> M&amp;A Intelligence</div>
    </div>
    <h1>Restricted Access</h1>
    <p class="sub">This tool is for authorised use only.<br>Enter the access password to continue.</p>
    <form method="POST" action="/auth">
      <label for="password">Access Password</label>
      <input type="password" id="password" name="password"
             placeholder="Enter password" autofocus required>
      <button type="submit">Continue →</button>
      <div id="pw-error" style="display:none">Incorrect password. Please try again.</div>
    </form>
  </div>
  <div class="footer">Not for distribution &nbsp;·&nbsp; Powered by GP Bullhound</div>
</body>
</html>"""


# ---------------------------------------------------------------------------
# Landing page HTML
# ---------------------------------------------------------------------------

LANDING_HTML = """<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>Strategic Fit Engine</title>
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link href="https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@300;400;600;700&display=swap" rel="stylesheet">
  <style>
    *{box-sizing:border-box;margin:0;padding:0}
    body{font-family:'IBM Plex Sans','Helvetica Neue',Arial,sans-serif;
      background:#fff;min-height:100vh;display:flex;flex-direction:column;color:#252850}

    /* ── Header ── */
    header{padding:0 56px;height:64px;display:flex;align-items:center;
      justify-content:space-between;border-bottom:2px solid #252850}
    .logo{font-size:15px;font-weight:700;color:#252850;letter-spacing:.02em}
    .logo span{color:#CC0605}
    .header-right{font-size:11px;color:#888;letter-spacing:.04em;text-transform:uppercase}

    /* ── Layout ── */
    main{flex:1;display:flex}
    .left-panel{width:340px;min-width:340px;background:#252850;padding:56px 48px;
      display:flex;flex-direction:column}
    .right-panel{flex:1;padding:56px 64px;overflow-y:auto}

    /* ── Left panel ── */
    .section-label{font-size:10px;font-weight:700;letter-spacing:.12em;
      text-transform:uppercase;color:#CC0605;margin-bottom:18px;
      display:flex;align-items:center;gap:8px}
    .section-label::before{content:'';display:inline-block;width:10px;height:10px;background:#CC0605}
    .panel-title{font-size:26px;font-weight:700;color:#fff;line-height:1.25;margin-bottom:20px}
    .panel-desc{font-size:13px;color:rgba(255,255,255,.55);line-height:1.7;margin-bottom:40px}

    .pipeline{display:flex;flex-direction:column;gap:0}
    .pipe-step{display:flex;align-items:flex-start;gap:16px;padding:16px 0;
      border-top:1px solid rgba(255,255,255,.1)}
    .pipe-step:last-child{border-bottom:1px solid rgba(255,255,255,.1)}
    .pipe-num{font-size:10px;font-weight:700;color:#CC0605;letter-spacing:.08em;
      text-transform:uppercase;min-width:28px;padding-top:2px}
    .pipe-name{font-size:13px;font-weight:600;color:#fff;margin-bottom:2px}
    .pipe-sub{font-size:11px;color:rgba(255,255,255,.4)}

    .panel-footer{margin-top:auto;padding-top:40px}
    .panel-footer a{font-size:12px;color:rgba(255,255,255,.35);text-decoration:none;
      transition:color .2s}
    .panel-footer a:hover{color:#CC0605}

    /* ── Right panel ── */
    .page-label{font-size:10px;font-weight:700;letter-spacing:.12em;text-transform:uppercase;
      color:#CC0605;display:flex;align-items:center;gap:8px;margin-bottom:16px}
    .page-label::before{content:'';display:inline-block;width:10px;height:10px;background:#CC0605}
    h1{font-size:32px;font-weight:700;color:#252850;line-height:1.2;margin-bottom:10px}
    .subtitle{font-size:14px;color:#666;line-height:1.65;margin-bottom:40px;max-width:560px}

    /* ── Presets ── */
    .presets-label{font-size:10px;font-weight:700;letter-spacing:.1em;text-transform:uppercase;
      color:#888;margin-bottom:12px}
    .presets{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:36px}
    .preset-btn{background:#fff;border:1px solid #d0d5e8;color:#252850;
      font-size:12px;font-weight:600;padding:7px 14px;cursor:pointer;
      font-family:inherit;transition:all .15s;letter-spacing:.01em}
    .preset-btn:hover{background:#252850;color:#fff;border-color:#252850}

    /* ── Form ── */
    .form-grid{display:grid;grid-template-columns:1fr 1fr;gap:24px;margin-bottom:24px}
    .field{display:flex;flex-direction:column}
    .field.full{grid-column:1/-1}
    label{font-size:10px;font-weight:700;letter-spacing:.1em;text-transform:uppercase;
      color:#252850;margin-bottom:8px}
    input{border:1px solid #d0d5e8;padding:12px 14px;font-size:14px;color:#252850;
      font-family:inherit;outline:none;transition:border .15s;background:#fff}
    input::placeholder{color:#aaa}
    input:focus{border-color:#252850}
    .hint{font-size:11px;color:#999;margin-top:5px;line-height:1.4}

    /* ── Submit ── */
    .submit-row{display:flex;align-items:center;gap:20px;margin-top:8px}
    .submit-btn{background:#CC0605;color:#fff;border:none;padding:14px 36px;
      font-size:14px;font-weight:700;font-family:inherit;cursor:pointer;
      letter-spacing:.04em;text-transform:uppercase;transition:background .15s}
    .submit-btn:hover{background:#a80504}
    .submit-btn:disabled{background:#ccc;cursor:not-allowed}

    /* ── Mode tabs ── */
    .mode-tabs{display:flex;gap:0;margin-bottom:28px;border:1px solid #d0d5e8}
    .mode-tab{flex:1;padding:11px 0;font-size:12px;font-weight:700;font-family:inherit;
      letter-spacing:.04em;text-transform:uppercase;cursor:pointer;border:none;
      background:#fff;color:#888;transition:all .15s;text-align:center}
    .mode-tab.active{background:#252850;color:#fff}
    .mode-tab:first-child{border-right:1px solid #d0d5e8}

    /* ── Footer ── */
    footer{padding:18px 56px;border-top:1px solid #e8eaf0;
      display:flex;justify-content:space-between;align-items:center}
    .footer-brand{font-size:13px;font-weight:700;color:#252850}
    .footer-brand span{color:#CC0605}
    .footer-right{font-size:11px;color:#aaa}
  </style>
</head>
<body>

<header>
  <div class="logo">Strategic Fit Engine &nbsp;<span>·</span>&nbsp; M&amp;A Intelligence</div>
  <div class="header-right" style="display:flex;align-items:center;gap:10px">
    <span style="font-size:11px;color:#aaa;letter-spacing:.04em;text-transform:uppercase">Powered by</span>
    <div style="background:#252850;padding:6px 14px;display:flex;align-items:center;gap:8px">
      <svg xmlns="http://www.w3.org/2000/svg" width="83.729" height="65.861" viewBox="0 0 83.729 65.861" style="height:24px;width:auto"><path d="M145.038,197c.1.1.2.207.306.309a.355.355,0,0,0,.5.094c.078-.034.158-.064.231-.094a.7.7,0,0,1,.3.449,1.8,1.8,0,0,0,.137.358c.113.232.245.455.35.689a.8.8,0,0,0,.523.459c.426.134.848.283,1.273.422a4.67,4.67,0,0,1,1.585.844c.574.481,1.142.968,1.713,1.452.3.251.593.5.886.752a1.421,1.421,0,0,1,.54.855c.028.186.012.378.022.566.013.274.016.55.051.822a.971.971,0,0,0,.575.775,2.5,2.5,0,0,0,.28.132,17.342,17.342,0,0,0,2.68.835,3.631,3.631,0,0,1,.695.192,1.027,1.027,0,0,1,.67,1.3c-.038.15-.094.295-.139.443a3.085,3.085,0,0,0-.09,1.253c.026.222.067.442.108.662a1.654,1.654,0,0,1-.025.716,5.265,5.265,0,0,1-.812,1.91,3.657,3.657,0,0,1-1.262,1.123,13.617,13.617,0,0,1-1.348.581,3.1,3.1,0,0,1-.658.123.8.8,0,0,0-.548.306,1.792,1.792,0,0,1-1.5.677,4.88,4.88,0,0,1-2.954-.92,1.292,1.292,0,0,0-.953-.235,3.979,3.979,0,0,0-.851.2,7.019,7.019,0,0,0-3.112,2.349,3.159,3.159,0,0,0-.673,1.944c-.006.774.044,1.549.075,2.322a13.282,13.282,0,0,1-.078,2.5,5.247,5.247,0,0,1-.188.8,19.087,19.087,0,0,1-4.387,7.4c-.554.589-1.135,1.154-1.7,1.729-.164.165-.33.328-.494.492a6.688,6.688,0,0,0-1.014,1.3c-1.725,2.869-3.468,5.728-5.2,8.6a42.319,42.319,0,0,0-2.144,4.148c-.435.944-.844,1.9-1.21,2.876a7.841,7.841,0,0,0-.512,3.01c.017.568-.012,1.137.01,1.7a7.676,7.676,0,0,0,.132,1.308,2.628,2.628,0,0,0,1.277,1.71c.278.171.567.327.837.511a1.388,1.388,0,0,1,.62,1.028,1.674,1.674,0,0,1-1.025,1.71,2.842,2.842,0,0,1-.879.18c-.947.018-1.894.005-2.841-.009a1.229,1.229,0,0,1-.823-.386,2.3,2.3,0,0,1-.556-.89,6.318,6.318,0,0,1-.35-2.1c-.011-.525.022-1.051-.005-1.575-.069-1.307-.157-2.612-.245-3.918-.047-.7-.106-1.39-.171-2.084a2.1,2.1,0,0,0-.126-.6,1.781,1.781,0,0,1,.241-1.78,9.679,9.679,0,0,0,1.594-3.688,8.327,8.327,0,0,0,.025-2.644c-.095-.673-.206-1.345-.263-2.021a10.35,10.35,0,0,1,0-1.471c.04-.678.117-1.355.176-2.032.007-.076,0-.152,0-.247-.162-.019-.322-.043-.482-.056a28.429,28.429,0,0,1-4.219-.712,57.123,57.123,0,0,1-5.665-1.75,21.981,21.981,0,0,0-2.692-.7,25.049,25.049,0,0,0-3.4-.47c-.9-.066-1.8-.133-2.706-.156a15.573,15.573,0,0,0-3.086.216,3.48,3.48,0,0,0-1.049.346,1.673,1.673,0,0,0-.831,1c-.315.964-.613,1.937-.987,2.878a13.815,13.815,0,0,1-3.263,4.908,31.925,31.925,0,0,1-6.736,5.008,10.825,10.825,0,0,1-1.026.484q-1.3.547-2.6,1.072a2.921,2.921,0,0,0-1.085.726,9.845,9.845,0,0,0-.826.991c-.5.718-.994,1.447-1.463,2.189-.45.713-.889,1.435-1.291,2.176-.422.778-.8,1.582-1.188,2.376a2.879,2.879,0,0,0-.149.412,1.709,1.709,0,0,0,.283,1.611c.206.276.435.534.643.809a1.5,1.5,0,0,1,.151,1.518.6.6,0,0,1-.374.372,3.43,3.43,0,0,1-.769.2c-1.081.081-2.165.138-3.248.2a2.546,2.546,0,0,1-.54-.038,1.046,1.046,0,0,1-.827-.628,1.965,1.965,0,0,1-.114-1.553q.158-.423.342-.835c.458-1.029.934-2.051,1.38-3.086q.878-2.041,1.918-4a13.764,13.764,0,0,0,1.27-3.451c.106-.469.265-.925.395-1.389a1.863,1.863,0,0,1,.8-1.041,8.944,8.944,0,0,0,.935-.728,4.048,4.048,0,0,0,1.194-1.895c.8-2.74,1.75-5.434,2.665-8.139.4-1.188.718-2.406,1.072-3.61a.235.235,0,0,0,0-.073c-.26.176-.508.362-.772.521a9.348,9.348,0,0,1-2.234.926,13.335,13.335,0,0,1-4.288.511,8.826,8.826,0,0,1-3.243-.777,3.714,3.714,0,0,1-.976-.623,2.49,2.49,0,0,1-.562-.708,1.759,1.759,0,0,1,.2-.054,7.485,7.485,0,0,1,1.8.029,15.821,15.821,0,0,0,2.239.134,8.979,8.979,0,0,0,7.616-4.409c.495-.8.938-1.625,1.39-2.446a15.758,15.758,0,0,1,3.708-4.5,5.175,5.175,0,0,1,.729-.531,19.466,19.466,0,0,1,5.258-2.153,17.969,17.969,0,0,1,2.6-.394c1.151-.1,2.3-.09,3.458-.086.843,0,1.686-.047,2.529-.094.592-.033,1.184-.09,1.774-.151a17.729,17.729,0,0,0,2.725-.566c1.6-.426,3.2-.875,4.81-1.282,1.475-.373,2.962-.7,4.442-1.054a43.05,43.05,0,0,0,7.265-2.4c.747-.326,1.527-.578,2.293-.863l.629-.231a6.817,6.817,0,0,0,3.239-2.423c.845-1.155,1.756-2.255,2.705-3.324a24.031,24.031,0,0,1,3.589-3.41l.416-.307a1.655,1.655,0,0,0,.6-.822,5.144,5.144,0,0,1,.555-1.19,4.973,4.973,0,0,1,1.624-1.525q.9-.549,1.809-1.1a.813.813,0,0,0,.11-.1Z" transform="translate(-74.671 -197)" fill="#fff"/></svg>
      <span style="font-size:12px;font-weight:700;color:#fff;letter-spacing:.02em;font-family:'IBM Plex Sans',sans-serif">GP Bullhound</span>
    </div>
  </div>
</header>

<main>
  <!-- Left panel -->
  <div class="left-panel">
    <div class="section-label">M&amp;A Intelligence</div>
    <div class="panel-title" id="lp-title">AI-Powered<br>Target Screening</div>
    <div class="panel-desc" id="lp-desc">
      Three inputs. A full analyst-grade report. Competitor threats, market signals,
      and acquisition readiness — all scored and ranked.
    </div>

    <div class="pipeline">
      <div class="pipe-step">
        <div class="pipe-num">01</div>
        <div>
          <div class="pipe-name" id="lp-s1-name">Buyer Intelligence</div>
          <div class="pipe-sub" id="lp-s1-sub">Strategy, dry powder &amp; M&amp;A thesis</div>
        </div>
      </div>
      <div class="pipe-step">
        <div class="pipe-num">02</div>
        <div>
          <div class="pipe-name" id="lp-s2-name">Target Longlist</div>
          <div class="pipe-sub" id="lp-s2-sub">Buyer-led discovery of all candidates</div>
        </div>
      </div>
      <div class="pipe-step">
        <div class="pipe-num">03</div>
        <div>
          <div class="pipe-name" id="lp-s3-name">Strategic Fit Scoring</div>
          <div class="pipe-sub">8 criteria, ranked and weighted</div>
        </div>
      </div>
      <div class="pipe-step">
        <div class="pipe-num">04</div>
        <div>
          <div class="pipe-name">Interactive Report</div>
          <div class="pipe-sub">HTML dashboard + PPTX deck</div>
        </div>
      </div>
    </div>

    <div class="panel-footer">
      <a href="/report" target="_blank">↗ View last report</a>
    </div>
  </div>

  <!-- Right panel -->
  <div class="right-panel">
    <div class="page-label">New Analysis</div>

    <!-- Mode tabs -->
    <div class="mode-tabs">
      <button class="mode-tab active" id="tab-buy" onclick="switchMode('buy')">
        Buy-Side Advisory &nbsp;·&nbsp; Find Targets
      </button>
      <button class="mode-tab" id="tab-sell" onclick="switchMode('sell')">
        Sell-Side Advisory &nbsp;·&nbsp; Find Acquirers
      </button>
    </div>

    <h1 id="rp-title">Define Your Screening Parameters</h1>
    <p class="subtitle" id="rp-subtitle">
      Enter a strategic buyer, target sector, and geography. The engine will map
      competitor acquisitions, scan for market signals, discover targets, and
      generate a ranked report in minutes.
    </p>

    <div class="presets-label">Quick Presets</div>
    <div class="presets" id="presets-buy">
      <button class="preset-btn" onclick="preset('Salesforce','European Healthcare Vertical SaaS','UK/Germany/France/Nordics/Netherlands')">Salesforce × EU Health</button>
      <button class="preset-btn" onclick="preset('Microsoft','European FinTech SaaS','UK/Germany/Netherlands/Sweden')">Microsoft × EU FinTech</button>
      <button class="preset-btn" onclick="preset('ServiceNow','European HR Tech SaaS','UK/Germany/France')">ServiceNow × EU HR Tech</button>
      <button class="preset-btn" onclick="preset('Workday','European Payroll &amp; Workforce SaaS','UK/Germany/Netherlands')">Workday × EU Payroll</button>
    </div>
    <div class="presets" id="presets-sell" style="display:none">
      <button class="preset-btn" onclick="preset('Attensi','Corporate EdTech SaaS','Europe')">Attensi · EdTech exit</button>
      <button class="preset-btn" onclick="preset('Phoebe AI','Healthcare AI SaaS','UK/Germany/France')">Phoebe AI · Healthcare exit</button>
      <button class="preset-btn" onclick="preset('Leapsome','HR Tech SaaS','Europe')">Leapsome · HR Tech exit</button>
      <button class="preset-btn" onclick="preset('Wayflyer','FinTech SaaS','UK/Ireland/Europe')">Wayflyer · FinTech exit</button>
    </div>

    <form method="POST" action="/run" id="analysis-form" onsubmit="handleSubmit(event)">
      <input type="hidden" name="mode" id="mode-input" value="buy">
      <div class="form-grid">
        <div class="field">
          <label id="company-label">Strategic Buyer</label>
          <input type="text" name="company" id="company"
            placeholder="e.g. Salesforce, Duolingo, Stripe" value="" required>
          <p class="hint" id="company-hint">The acquiring company. Scoring criteria will be derived from their competitive gaps.</p>
        </div>
        <div class="field">
          <label id="sector-label">Target Sector</label>
          <input type="text" name="sector" id="sector"
            placeholder="e.g. Healthcare Vertical SaaS, EdTech, FinTech payments"
            value="" required>
          <p class="hint" id="sector-hint">Be specific — vertical SaaS, FinTech payments, HR Tech, etc.</p>
        </div>
        <div class="field full">
          <label id="geo-label">Geography / Region</label>
          <input type="text" name="geography" id="geography"
            placeholder="e.g. UK/Germany/France/Nordics/Netherlands"
            value="" required>
          <p class="hint" id="geo-hint">Countries or regions to screen. Separate with / or commas.</p>
        </div>
      </div>
      <div class="submit-row">
        <button type="submit" class="submit-btn" id="submit-btn">Run Analysis →</button>
      </div>
    </form>
  </div>
</main>

<footer>
  <div class="footer-brand">Strategic Fit Engine <span>·</span> M&amp;A Intelligence</div>
  <div class="footer-right">AI-powered screening · Not for distribution</div>
</footer>

<script>
var _currentMode = 'buy';

var _copy = {
  buy: {
    title: 'Define Your Screening Parameters',
    subtitle: 'Enter a strategic buyer, target sector, and geography. The engine will map competitor acquisitions, scan for market signals, discover targets, and generate a ranked report in minutes.',
    companyLabel: 'Strategic Buyer',
    companyHint: 'The acquiring company. Scoring criteria will be derived from their competitive gaps.',
    companyPlaceholder: 'e.g. Salesforce, Duolingo, Stripe',
    sectorLabel: 'Target Sector',
    sectorHint: 'Be specific — vertical SaaS, FinTech payments, HR Tech, etc.',
    geoLabel: 'Geography / Region',
    geoHint: 'Countries or regions to screen. Separate with / or commas.',
    btnText: 'Run Buy-Side Analysis →',
    lpTitle: 'AI-Powered<br>Target Screening',
    lpDesc: 'Three inputs. A full analyst-grade report. Competitor threats, market signals, and acquisition readiness — all scored and ranked.',
    s1Name: 'Buyer Intelligence', s1Sub: 'Strategy, dry powder & M\u0026A thesis',
    s2Name: 'Target Longlist',    s2Sub: 'Buyer-led discovery of all candidates',
    s3Name: 'Strategic Fit Scoring',
  },
  sell: {
    title: 'Define Your Exit Parameters',
    subtitle: 'Enter the seller company, sector, and geography of potential acquirers. The engine will identify strategic and PE buyers most likely to pay a premium — scored, ranked, and ready to pitch.',
    companyLabel: 'Seller Company',
    companyHint: 'The company seeking an exit. The engine will profile its appeal and identify ideal acquirers.',
    companyPlaceholder: 'e.g. Attensi, Leapsome, Wayflyer',
    sectorLabel: 'Sector',
    sectorHint: 'The seller sector — used to find relevant strategic and PE acquirers.',
    geoLabel: 'Acquirer Geography',
    geoHint: 'Where to look for potential acquirers. Separate with / or commas.',
    btnText: 'Run Sell-Side Analysis →',
    lpTitle: 'AI-Powered<br>Exit Strategy',
    lpDesc: 'Three inputs. A full sell-side brief. Acquirer appetite signals, strategic fit, and premium rationale — all scored and ranked.',
    s1Name: 'Seller Profile',    s1Sub: 'Appeal analysis and acquirer criteria',
    s2Name: 'Acquirer Longlist', s2Sub: 'Strategic & PE acquirers identified',
    s3Name: 'Acquirer Fit Scoring',
  }
};

function switchMode(mode) {
  _currentMode = mode;
  var c = _copy[mode];
  document.getElementById('mode-input').value = mode;
  document.getElementById('tab-buy').className  = 'mode-tab' + (mode === 'buy'  ? ' active' : '');
  document.getElementById('tab-sell').className = 'mode-tab' + (mode === 'sell' ? ' active' : '');
  document.getElementById('presets-buy').style.display  = (mode === 'buy')  ? '' : 'none';
  document.getElementById('presets-sell').style.display = (mode === 'sell') ? '' : 'none';
  document.getElementById('rp-title').textContent    = c.title;
  document.getElementById('rp-subtitle').textContent = c.subtitle;
  document.getElementById('company-label').textContent = c.companyLabel;
  document.getElementById('company-hint').textContent  = c.companyHint;
  document.getElementById('company').placeholder = c.companyPlaceholder;
  document.getElementById('sector-label').textContent = c.sectorLabel;
  document.getElementById('sector-hint').textContent  = c.sectorHint;
  document.getElementById('geo-label').textContent = c.geoLabel;
  document.getElementById('geo-hint').textContent  = c.geoHint;
  document.getElementById('submit-btn').textContent = c.btnText;
  document.getElementById('lp-title').innerHTML  = c.lpTitle;
  document.getElementById('lp-desc').textContent = c.lpDesc;
  document.getElementById('lp-s1-name').textContent = c.s1Name;
  document.getElementById('lp-s1-sub').textContent  = c.s1Sub;
  document.getElementById('lp-s2-name').textContent = c.s2Name;
  document.getElementById('lp-s2-sub').textContent  = c.s2Sub;
  document.getElementById('lp-s3-name').textContent = c.s3Name;
}

function preset(company, sector, geo) {
  document.getElementById('company').value   = company;
  document.getElementById('sector').value    = sector;
  document.getElementById('geography').value = geo;
}
function handleSubmit(e) {
  const btn = document.getElementById('submit-btn');
  btn.disabled = true;
  btn.textContent = 'Starting...';
}
</script>
</body>
</html>"""


# ---------------------------------------------------------------------------
# Progress page HTML (returned after form submit)
# ---------------------------------------------------------------------------

def _progress_html(company: str, sector: str, geography: str, mode: str = "buy") -> str:
    is_sell = (mode == "sell")
    param_label = "Seller"            if is_sell else "Buyer"
    s1_name     = "Seller Profile"    if is_sell else "Buyer Intelligence"
    s1_sub      = "M&amp;A appeal &amp; acquirer criteria" if is_sell else "Strategy, dry powder &amp; M&amp;A thesis"
    s2_name     = "Acquirer Longlist" if is_sell else "Target Longlist"
    s2_sub      = "Strategic &amp; PE acquirers" if is_sell else "Buyer-led discovery of all candidates"
    s3_name     = "Acquirer Fit Scoring" if is_sell else "Strategic Fit Scoring"
    js_labels   = ("['', 'Seller Profile', 'Acquirer Longlist', 'Acquirer Fit Scoring', 'Generating Report']"
                   if is_sell else
                   "['', 'Buyer Intelligence', 'Target Longlist', 'Scoring & Ranking', 'Generating Report']")
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>Running Analysis — Strategic Fit Engine</title>
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link href="https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@300;400;600;700&display=swap" rel="stylesheet">
  <style>
    *{{box-sizing:border-box;margin:0;padding:0}}
    body{{font-family:'IBM Plex Sans','Helvetica Neue',Arial,sans-serif;
      background:#fff;min-height:100vh;display:flex;flex-direction:column;color:#252850}}

    /* ── Header ── */
    header{{padding:0 56px;height:64px;display:flex;align-items:center;
      justify-content:space-between;border-bottom:2px solid #252850}}
    .logo{{font-size:15px;font-weight:700;color:#252850;letter-spacing:.02em}}
    .logo span{{color:#CC0605}}
    .header-right{{font-size:11px;color:#888;letter-spacing:.04em;text-transform:uppercase}}
    .header-right a{{color:#252850;text-decoration:none;font-weight:600;
      font-size:12px;letter-spacing:.04em}}
    .header-right a:hover{{color:#CC0605}}

    /* ── Layout ── */
    main{{flex:1;display:flex}}
    .left-panel{{width:340px;min-width:340px;background:#252850;padding:56px 48px;
      display:flex;flex-direction:column}}
    .right-panel{{flex:1;padding:56px 64px;overflow-y:auto}}

    /* ── Left panel ── */
    .section-label{{font-size:10px;font-weight:700;letter-spacing:.12em;
      text-transform:uppercase;color:#CC0605;margin-bottom:18px;
      display:flex;align-items:center;gap:8px}}
    .section-label::before{{content:'';display:inline-block;width:10px;height:10px;background:#CC0605}}
    .panel-title{{font-size:26px;font-weight:700;color:#fff;line-height:1.25;margin-bottom:16px}}
    .params-block{{margin-bottom:36px}}
    .param-row{{display:flex;flex-direction:column;margin-bottom:14px}}
    .param-key{{font-size:10px;font-weight:700;letter-spacing:.1em;text-transform:uppercase;
      color:rgba(255,255,255,.35);margin-bottom:3px}}
    .param-val{{font-size:13px;font-weight:600;color:#fff}}

    .pipeline{{display:flex;flex-direction:column;gap:0}}
    .pipe-step{{display:flex;align-items:flex-start;gap:16px;padding:14px 0;
      border-top:1px solid rgba(255,255,255,.1)}}
    .pipe-step:last-child{{border-bottom:1px solid rgba(255,255,255,.1)}}
    .pipe-num{{font-size:10px;font-weight:700;color:rgba(255,255,255,.3);letter-spacing:.08em;
      text-transform:uppercase;min-width:28px;padding-top:2px;transition:color .3s}}
    .pipe-step.active .pipe-num{{color:#CC0605}}
    .pipe-step.complete .pipe-num{{color:#4CAF50}}
    .pipe-name{{font-size:13px;font-weight:600;color:rgba(255,255,255,.4);margin-bottom:2px;transition:color .3s}}
    .pipe-step.active .pipe-name{{color:#fff}}
    .pipe-step.complete .pipe-name{{color:rgba(255,255,255,.7)}}
    .pipe-sub{{font-size:11px;color:rgba(255,255,255,.25)}}
    .pipe-icon{{margin-left:auto;font-size:14px;color:rgba(255,255,255,.2);transition:color .3s}}
    .pipe-step.active .pipe-icon{{color:#CC0605}}
    .pipe-step.complete .pipe-icon{{color:#4CAF50}}

    /* ── Right panel ── */
    .page-label{{font-size:10px;font-weight:700;letter-spacing:.12em;text-transform:uppercase;
      color:#CC0605;display:flex;align-items:center;gap:8px;margin-bottom:16px}}
    .page-label::before{{content:'';display:inline-block;width:10px;height:10px;background:#CC0605}}
    h2{{font-size:28px;font-weight:700;color:#252850;line-height:1.2;margin-bottom:8px}}
    .running-sub{{font-size:13px;color:#888;margin-bottom:36px}}

    /* ── Progress bar ── */
    .progress-wrap{{margin-bottom:36px}}
    .progress-meta{{display:flex;justify-content:space-between;margin-bottom:8px}}
    .progress-label{{font-size:11px;font-weight:700;letter-spacing:.08em;
      text-transform:uppercase;color:#888}}
    .progress-pct{{font-size:11px;font-weight:700;color:#252850}}
    .progress-track{{background:#f0f2f7;height:4px;overflow:hidden}}
    .progress-fill{{height:4px;background:#CC0605;width:0%;transition:width .6s ease}}

    /* ── Log ── */
    .log-label{{font-size:10px;font-weight:700;letter-spacing:.1em;text-transform:uppercase;
      color:#888;margin-bottom:10px}}
    .log{{background:#f7f8fc;border:1px solid #e8eaf0;padding:20px;
      height:300px;overflow-y:auto;font-family:'IBM Plex Mono','Courier New',monospace;
      font-size:12px;line-height:1.8}}
    .log .msg-log{{color:#666}}
    .log .msg-info{{color:#444}}
    .log .msg-step{{color:#252850;font-weight:700}}
    .log .msg-done{{color:#2E7D32;font-weight:600}}
    .log .msg-error{{color:#CC0605}}
    .log .msg-complete{{color:#252850;font-weight:700;font-size:13px}}

    /* ── Result ── */
    .result-box{{margin-top:28px;padding:28px 32px;border-left:4px solid #2E7D32;
      background:#f4faf4;display:none}}
    .result-box.show{{display:block}}
    .result-title{{font-size:16px;font-weight:700;color:#252850;margin-bottom:4px}}
    .result-sub{{font-size:13px;color:#666;margin-bottom:20px}}
    .open-btn{{display:inline-block;padding:13px 32px;background:#CC0605;color:#fff;
      font-size:13px;font-weight:700;font-family:inherit;text-decoration:none;
      letter-spacing:.04em;text-transform:uppercase;transition:background .15s;
      margin-right:12px}}
    .open-btn:hover{{background:#a80504}}
    .new-btn{{display:inline-block;padding:13px 24px;background:#fff;color:#252850;
      font-size:13px;font-weight:600;font-family:inherit;text-decoration:none;
      border:1px solid #d0d5e8;transition:all .15s}}
    .new-btn:hover{{background:#252850;color:#fff;border-color:#252850}}

    /* ── Error ── */
    .error-box{{margin-top:28px;padding:24px 28px;border-left:4px solid #CC0605;
      background:#fff5f5;display:none}}
    .error-box.show{{display:block}}
    .error-title{{font-size:15px;font-weight:700;color:#CC0605;margin-bottom:8px}}
    .error-msg{{font-size:13px;color:#444;margin-bottom:16px;font-family:'IBM Plex Mono','Courier New',monospace}}
    .retry-link{{font-size:13px;color:#252850;text-decoration:none;font-weight:600;
      border-bottom:1px solid #252850}}
    .retry-link:hover{{color:#CC0605;border-color:#CC0605}}

    /* ── Spinner ── */
    @keyframes spin{{to{{transform:rotate(360deg)}}}}
    .spinner{{display:inline-block;width:14px;height:14px;
      border:2px solid rgba(204,6,5,.2);border-top-color:#CC0605;
      border-radius:50%;animation:spin .8s linear infinite;vertical-align:middle;
      margin-right:8px}}
    @keyframes pulse{{0%,100%{{opacity:1}}50%{{opacity:.4}}}}
    .pulsing{{animation:pulse 1.8s ease-in-out infinite}}

    /* ── Footer ── */
    footer{{padding:18px 56px;border-top:1px solid #e8eaf0;
      display:flex;justify-content:space-between;align-items:center}}
    .footer-brand{{font-size:13px;font-weight:700;color:#252850}}
    .footer-brand span{{color:#CC0605}}
    .footer-right{{font-size:11px;color:#aaa}}
  </style>
</head>
<body>

<header>
  <div class="logo">Strategic Fit Engine &nbsp;<span>·</span>&nbsp; M&amp;A Intelligence</div>
  <div class="header-right" style="display:flex;align-items:center;gap:12px">
    <a href="/" style="font-size:12px;color:#888;letter-spacing:.04em;text-decoration:none;font-weight:600">← New Analysis</a>
    <div style="background:#252850;padding:6px 14px;display:flex;align-items:center;gap:8px">
      <svg xmlns="http://www.w3.org/2000/svg" width="83.729" height="65.861" viewBox="0 0 83.729 65.861" style="height:24px;width:auto"><path d="M145.038,197c.1.1.2.207.306.309a.355.355,0,0,0,.5.094c.078-.034.158-.064.231-.094a.7.7,0,0,1,.3.449,1.8,1.8,0,0,0,.137.358c.113.232.245.455.35.689a.8.8,0,0,0,.523.459c.426.134.848.283,1.273.422a4.67,4.67,0,0,1,1.585.844c.574.481,1.142.968,1.713,1.452.3.251.593.5.886.752a1.421,1.421,0,0,1,.54.855c.028.186.012.378.022.566.013.274.016.55.051.822a.971.971,0,0,0,.575.775,2.5,2.5,0,0,0,.28.132,17.342,17.342,0,0,0,2.68.835,3.631,3.631,0,0,1,.695.192,1.027,1.027,0,0,1,.67,1.3c-.038.15-.094.295-.139.443a3.085,3.085,0,0,0-.09,1.253c.026.222.067.442.108.662a1.654,1.654,0,0,1-.025.716,5.265,5.265,0,0,1-.812,1.91,3.657,3.657,0,0,1-1.262,1.123,13.617,13.617,0,0,1-1.348.581,3.1,3.1,0,0,1-.658.123.8.8,0,0,0-.548.306,1.792,1.792,0,0,1-1.5.677,4.88,4.88,0,0,1-2.954-.92,1.292,1.292,0,0,0-.953-.235,3.979,3.979,0,0,0-.851.2,7.019,7.019,0,0,0-3.112,2.349,3.159,3.159,0,0,0-.673,1.944c-.006.774.044,1.549.075,2.322a13.282,13.282,0,0,1-.078,2.5,5.247,5.247,0,0,1-.188.8,19.087,19.087,0,0,1-4.387,7.4c-.554.589-1.135,1.154-1.7,1.729-.164.165-.33.328-.494.492a6.688,6.688,0,0,0-1.014,1.3c-1.725,2.869-3.468,5.728-5.2,8.6a42.319,42.319,0,0,0-2.144,4.148c-.435.944-.844,1.9-1.21,2.876a7.841,7.841,0,0,0-.512,3.01c.017.568-.012,1.137.01,1.7a7.676,7.676,0,0,0,.132,1.308,2.628,2.628,0,0,0,1.277,1.71c.278.171.567.327.837.511a1.388,1.388,0,0,1,.62,1.028,1.674,1.674,0,0,1-1.025,1.71,2.842,2.842,0,0,1-.879.18c-.947.018-1.894.005-2.841-.009a1.229,1.229,0,0,1-.823-.386,2.3,2.3,0,0,1-.556-.89,6.318,6.318,0,0,1-.35-2.1c-.011-.525.022-1.051-.005-1.575-.069-1.307-.157-2.612-.245-3.918-.047-.7-.106-1.39-.171-2.084a2.1,2.1,0,0,0-.126-.6,1.781,1.781,0,0,1,.241-1.78,9.679,9.679,0,0,0,1.594-3.688,8.327,8.327,0,0,0,.025-2.644c-.095-.673-.206-1.345-.263-2.021a10.35,10.35,0,0,1,0-1.471c.04-.678.117-1.355.176-2.032.007-.076,0-.152,0-.247-.162-.019-.322-.043-.482-.056a28.429,28.429,0,0,1-4.219-.712,57.123,57.123,0,0,1-5.665-1.75,21.981,21.981,0,0,0-2.692-.7,25.049,25.049,0,0,0-3.4-.47c-.9-.066-1.8-.133-2.706-.156a15.573,15.573,0,0,0-3.086.216,3.48,3.48,0,0,0-1.049.346,1.673,1.673,0,0,0-.831,1c-.315.964-.613,1.937-.987,2.878a13.815,13.815,0,0,1-3.263,4.908,31.925,31.925,0,0,1-6.736,5.008,10.825,10.825,0,0,1-1.026.484q-1.3.547-2.6,1.072a2.921,2.921,0,0,0-1.085.726,9.845,9.845,0,0,0-.826.991c-.5.718-.994,1.447-1.463,2.189-.45.713-.889,1.435-1.291,2.176-.422.778-.8,1.582-1.188,2.376a2.879,2.879,0,0,0-.149.412,1.709,1.709,0,0,0,.283,1.611c.206.276.435.534.643.809a1.5,1.5,0,0,1,.151,1.518.6.6,0,0,1-.374.372,3.43,3.43,0,0,1-.769.2c-1.081.081-2.165.138-3.248.2a2.546,2.546,0,0,1-.54-.038,1.046,1.046,0,0,1-.827-.628,1.965,1.965,0,0,1-.114-1.553q.158-.423.342-.835c.458-1.029.934-2.051,1.38-3.086q.878-2.041,1.918-4a13.764,13.764,0,0,0,1.27-3.451c.106-.469.265-.925.395-1.389a1.863,1.863,0,0,1,.8-1.041,8.944,8.944,0,0,0,.935-.728,4.048,4.048,0,0,0,1.194-1.895c.8-2.74,1.75-5.434,2.665-8.139.4-1.188.718-2.406,1.072-3.61a.235.235,0,0,0,0-.073c-.26.176-.508.362-.772.521a9.348,9.348,0,0,1-2.234.926,13.335,13.335,0,0,1-4.288.511,8.826,8.826,0,0,1-3.243-.777,3.714,3.714,0,0,1-.976-.623,2.49,2.49,0,0,1-.562-.708,1.759,1.759,0,0,1,.2-.054,7.485,7.485,0,0,1,1.8.029,15.821,15.821,0,0,0,2.239.134,8.979,8.979,0,0,0,7.616-4.409c.495-.8.938-1.625,1.39-2.446a15.758,15.758,0,0,1,3.708-4.5,5.175,5.175,0,0,1,.729-.531,19.466,19.466,0,0,1,5.258-2.153,17.969,17.969,0,0,1,2.6-.394c1.151-.1,2.3-.09,3.458-.086.843,0,1.686-.047,2.529-.094.592-.033,1.184-.09,1.774-.151a17.729,17.729,0,0,0,2.725-.566c1.6-.426,3.2-.875,4.81-1.282,1.475-.373,2.962-.7,4.442-1.054a43.05,43.05,0,0,0,7.265-2.4c.747-.326,1.527-.578,2.293-.863l.629-.231a6.817,6.817,0,0,0,3.239-2.423c.845-1.155,1.756-2.255,2.705-3.324a24.031,24.031,0,0,1,3.589-3.41l.416-.307a1.655,1.655,0,0,0,.6-.822,5.144,5.144,0,0,1,.555-1.19,4.973,4.973,0,0,1,1.624-1.525q.9-.549,1.809-1.1a.813.813,0,0,0,.11-.1Z" transform="translate(-74.671 -197)" fill="#fff"/></svg>
      <span style="font-size:12px;font-weight:700;color:#fff;letter-spacing:.02em;font-family:'IBM Plex Sans',sans-serif">GP Bullhound</span>
    </div>
  </div>
</header>

<main>
  <!-- Left panel -->
  <div class="left-panel">
    <div class="section-label">Pipeline</div>
    <div class="panel-title">Analysis<br>In Progress</div>

    <div class="params-block">
      <div class="param-row">
        <div class="param-key">{param_label}</div>
        <div class="param-val">{_h(company)}</div>
      </div>
      <div class="param-row">
        <div class="param-key">Sector</div>
        <div class="param-val">{_h(sector)}</div>
      </div>
      <div class="param-row">
        <div class="param-key">Geography</div>
        <div class="param-val">{_h(geography)}</div>
      </div>
    </div>

    <div class="pipeline">
      <div class="pipe-step" id="step1">
        <div class="pipe-num">01</div>
        <div style="flex:1">
          <div class="pipe-name">{s1_name}</div>
          <div class="pipe-sub">{s1_sub}</div>
        </div>
        <div class="pipe-icon" id="icon1">○</div>
      </div>
      <div class="pipe-step" id="step2">
        <div class="pipe-num">02</div>
        <div style="flex:1">
          <div class="pipe-name">{s2_name}</div>
          <div class="pipe-sub">{s2_sub}</div>
        </div>
        <div class="pipe-icon" id="icon2">○</div>
      </div>
      <div class="pipe-step" id="step3">
        <div class="pipe-num">03</div>
        <div style="flex:1">
          <div class="pipe-name">{s3_name}</div>
          <div class="pipe-sub">8 criteria, ranked &amp; weighted</div>
        </div>
        <div class="pipe-icon" id="icon3">○</div>
      </div>
      <div class="pipe-step" id="step4">
        <div class="pipe-num">04</div>
        <div style="flex:1">
          <div class="pipe-name">Report Generation</div>
          <div class="pipe-sub">HTML dashboard + PPTX deck</div>
        </div>
        <div class="pipe-icon" id="icon4">○</div>
      </div>
    </div>
  </div>

  <!-- Right panel -->
  <div class="right-panel">
    <div class="page-label">Live Analysis</div>
    <h2><span class="spinner pulsing"></span>Running Pipeline</h2>
    <div class="running-sub">This typically takes 3–5 minutes. Do not close this tab.</div>

    <div class="progress-wrap">
      <div class="progress-meta">
        <div class="progress-label" id="progress-label">Initialising</div>
        <div class="progress-pct" id="progress-pct">0%</div>
      </div>
      <div class="progress-track">
        <div class="progress-fill" id="progress-fill"></div>
      </div>
    </div>

    <div class="log-label">Live Output</div>
    <div class="log" id="log"></div>

    <div class="result-box" id="result-box">
      <div class="result-title">✓ Analysis Complete</div>
      <div class="result-sub">Your report is ready. It will open automatically in a new tab.</div>
      <a href="/report" target="_blank" class="open-btn">Open Report →</a>
      <a href="/" class="new-btn">Run Another</a>
    </div>

    <div class="error-box" id="error-box">
      <div class="error-title">✗ Analysis Failed</div>
      <div id="error-msg" class="error-msg"></div>
      <a href="/" class="retry-link">← Try again</a>
    </div>
  </div>
</main>

<footer>
  <div class="footer-brand">Strategic Fit Engine <span>·</span> M&amp;A Intelligence</div>
  <div class="footer-right">AI-powered screening · Not for distribution</div>
</footer>

<script>
const log = document.getElementById('log');
const progressFill = document.getElementById('progress-fill');
const progressPct = document.getElementById('progress-pct');
const progressLabel = document.getElementById('progress-label');

const stepLabels = {js_labels};

function appendLog(msg) {{
  const div = document.createElement('div');
  div.className = 'msg-' + msg.type;
  div.textContent = msg.text;
  log.appendChild(div);
  log.scrollTop = log.scrollHeight;
}}

function setStep(n) {{
  [1,2,3,4].forEach(i => {{
    const el = document.getElementById('step' + i);
    const icon = document.getElementById('icon' + i);
    if (i < n) {{
      el.className = 'pipe-step complete';
      icon.textContent = '✓';
    }} else if (i === n) {{
      el.className = 'pipe-step active';
      icon.textContent = '⟳';
    }} else {{
      el.className = 'pipe-step';
      icon.textContent = '○';
    }}
  }});
  const pct = Math.round((n - 1) / 4 * 100);
  progressFill.style.width = pct + '%';
  progressPct.textContent = pct + '%';
  if (n <= 4) progressLabel.textContent = stepLabels[n];
}}

const src = new EventSource('/stream');
src.onmessage = function(e) {{
  const msg = JSON.parse(e.data);

  if (msg.type === 'end') {{
    src.close();
    if (msg.error) {{
      document.getElementById('error-msg').textContent = msg.error;
      document.getElementById('error-box').classList.add('show');
      [1,2,3,4].forEach(i => {{
        document.getElementById('step'+i).classList.remove('active');
      }});
    }} else {{
      progressFill.style.width = '100%';
      progressPct.textContent = '100%';
      progressLabel.textContent = 'Complete';
      [1,2,3,4].forEach(i => {{
        document.getElementById('step'+i).className = 'pipe-step complete';
        document.getElementById('icon'+i).textContent = '✓';
      }});
      document.getElementById('result-box').classList.add('show');
      // Auto-open report after brief delay
      setTimeout(() => window.open('/report', '_blank'), 1200);
    }}
    return;
  }}

  if (msg.step && msg.step > 0) setStep(msg.step);
  appendLog(msg);
}};

src.onerror = function() {{
  src.close();
  appendLog({{type:'error', text:'Connection to server lost. Check terminal for errors.'}});
}};
</script>
</body>
</html>"""


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    print("\n" + "="*60)
    print("  Strategic Fit Engine — Web Interface")
    print("="*60)
    print(f"\n  Open http://localhost:{port} in your browser\n")
    print("  Press Ctrl+C to stop\n")
    app.run(host="0.0.0.0", port=port, debug=False, threaded=True)
